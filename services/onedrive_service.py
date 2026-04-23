"""
OneDrive Service Module - Unified Microsoft Graph API Client
Integrates file operations, email, OneNote, Teams and other functionalities
"""
import io
from openpyxl import load_workbook
import requests
import json
import os
import base64
import time
import html2text
import bs4
from email import policy
from email.parser import BytesParser
from typing import Optional, Dict, Any, List, Generator
import urllib3
import sys
from datetime import datetime

from .mongo_service import MongoTokenService
from exceptions import MongoDBError

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def log(message: str, level: str = "INFO"):
    """Unified logging function"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    print(f"[{timestamp}] [{level}] [OneDrive] {message}", file=sys.stderr)


# HTML processing utility classes
class Mail2Text(BytesParser):
    """Email HTML to text converter"""
    def __init__(self, html_email_bytes):
        super().__init__(policy=policy.default)
        msg = self.parsebytes(html_email_bytes)
        self.text = msg.get_body(preferencelist=('plain', 'html'))
        if self.text is None:
            self.text = msg.get_body(preferencelist=('html',)).get_content()
        else:
            self.text = self.text.get_content()


class BeautifulSoup(bs4.BeautifulSoup):
    """Custom BeautifulSoup with optimized HTML processing"""
    def __init__(self, html_content):
        super().__init__(html_content, 'html.parser')
        for br in self.find_all('br'): 
            br.replace_with('\n')
        for li in self.find_all('li'): 
            li.insert_before('• ')

    def get_text(self):
        return super().get_text(separator='\n', strip=True)


class HTML2Text(html2text.HTML2Text):
    """Custom HTML2Text converter"""
    def __init__(self):
        super().__init__()
        self.body_width = 0
        self.ul_item_mark = '-'
        self.emphasis_mark = '*'
        self.wrap_links = False


# Data container classes
class Dir:
    """Directory data container"""
    def __init__(self, json_data):
        self.json_data = json_data
    
    def __getitem__(self, index):
        return Dict(self.json_data['value'][index])

    def __repr__(self):
        return 'Dir{0}'.format([i['name'] for i in self.json_data['value']])


class Dict(dict):
    """File/folder data container"""
    def __repr__(self):
        return 'Dict({0})'.format(self['name'])


# Teams related classes
class Base:
    """Teams base class"""
    def __init__(self, value, onedrive):
        self.value = value
        self.onedrive = onedrive

    def is_less_days_by_now(self, date_str, days=0.5):
        """Check if date is within specified days"""
        date_time = time.strptime(date_str.split('.')[0].rstrip('Z'), '%Y-%m-%dT%H:%M:%S')
        if date_time.tm_year >= 2020:
            return (time.time() - time.mktime(date_time)) < days * 24 * 3600
        else: 
            return False


class Chat(Base):
    """Teams chat class"""
    def is_less_days_by_now(self, days=0.5):
        return super().is_less_days_by_now(self.value['viewpoint']['lastMessageReadDateTime'], days)

    def read_messages(self):
        chat_id = self.value['id']
        self.messages = self.onedrive.call_rest_api(f'/chats/{chat_id}/messages', lambda v,o: Message(v,o))
        return self.messages


class Message(Base):
    """Teams message class"""
    def is_less_days_by_now(self, days=0.5):
        return super().is_less_days_by_now(self.value['lastModifiedDateTime'], days)

    def read_content(self):
        return self.value['body']['content']


# Main OneDrive service class
class OneDriverService:
    """
    Unified Microsoft Graph API Client
    Integrates file operations, email, OneNote, Teams and all other functionalities
    """
    BASE_URL = os.environ.get("BASE_URL")
    def __init__(self, token: str):
        """
        Initialize OneDrive service
        
        Args:
            token: unique_token (used directly for backend authentication)
        """
        log(f"==================== 初始化 OneDriverService ====================")
        log(f"收到 token: {token[:20]}...{token[-10:]}")
        log(f"Token 长度: {len(token)}")
        log(f"BASE_URL: {self.BASE_URL}")
        
        self.token = token
        self.headers = {'Authorization': f'Bearer {self.token}'}
        self.requests = requests
        
        log(f"认证头已设置: Authorization: Bearer {token[:20]}...{token[-10:]}")
        log(f"初始化完成")
        log(f"=================================================================")
    
    def _make_request(self, method: str, url: str, **kwargs) -> requests.Response:
        """Unified request method with detailed logging"""
        request_id = f"{int(time.time() * 1000)}"
        
        log(f"-------------------- 请求开始 [{request_id}] --------------------")
        log(f"请求方法: {method}")
        log(f"请求URL: {url}")
        log(f"请求头: {kwargs.get('headers', {})}")
        
        if 'params' in kwargs and kwargs['params']:
            log(f"查询参数: {json.dumps(kwargs['params'], ensure_ascii=False)}")
        
        if 'json' in kwargs and kwargs['json']:
            log(f"JSON数据: {json.dumps(kwargs['json'], ensure_ascii=False, indent=2)[:500]}...")
        
        if 'data' in kwargs and kwargs['data']:
            data_preview = str(kwargs['data'])[:200]
            log(f"请求体数据 (前200字符): {data_preview}...")
        
        kwargs.setdefault('headers', {}).update(self.headers)
        kwargs.setdefault('verify', False)
        
        log(f"最终请求头: {json.dumps(dict(kwargs['headers']), ensure_ascii=False)}")
        
        try:
            start_time = time.time()
            log(f"发送请求...")
            
            response = getattr(requests, method.lower())(url, **kwargs)
            
            elapsed_time = (time.time() - start_time) * 1000
            log(f"收到响应 - 耗时: {elapsed_time:.2f}ms")
            log(f"响应状态码: {response.status_code}")
            log(f"响应头: {dict(response.headers)}")
            
            # Try to parse response content
            content_type = response.headers.get('Content-Type', '')
            if 'application/json' in content_type:
                try:
                    response_json = response.json()
                    log(f"响应JSON数据 (前500字符): {json.dumps(response_json, ensure_ascii=False, indent=2)[:500]}...")
                    
                    # Log error details if present
                    if 'error' in response_json:
                        log(f"❌ 响应包含错误: {json.dumps(response_json['error'], ensure_ascii=False, indent=2)}", "ERROR")
                except json.JSONDecodeError as e:
                    log(f"⚠️ 无法解析JSON响应: {e}", "WARN")
                    log(f"响应文本 (前500字符): {response.text[:500]}...", "WARN")
            else:
                log(f"响应内容类型: {content_type}")
                if len(response.content) < 1000:
                    log(f"响应内容: {response.text[:500]}...")
                else:
                    log(f"响应内容大小: {len(response.content)} bytes")
            
            response.raise_for_status()
            
            log(f"✅ 请求成功 [{request_id}]")
            log(f"---------------------------------------------------------------\n")
            
            return response
            
        except requests.HTTPError as e:
            log(f"❌ HTTP错误 [{request_id}]: {e}", "ERROR")
            log(f"状态码: {e.response.status_code}", "ERROR")
            
            try:
                error_detail = e.response.json()
                log(f"错误详情: {json.dumps(error_detail, ensure_ascii=False, indent=2)}", "ERROR")
            except:
                log(f"错误响应文本: {e.response.text[:500]}...", "ERROR")
            
            log(f"---------------------------------------------------------------\n")
            raise
            
        except requests.ConnectionError as e:
            log(f"❌ 连接错误 [{request_id}]: {e}", "ERROR")
            log(f"无法连接到: {url}", "ERROR")
            log(f"---------------------------------------------------------------\n")
            raise
            
        except requests.Timeout as e:
            log(f"❌ 请求超时 [{request_id}]: {e}", "ERROR")
            log(f"---------------------------------------------------------------\n")
            raise
            
        except requests.RequestException as e:
            log(f"❌ 请求异常 [{request_id}]: {e}", "ERROR")
            log(f"异常类型: {type(e).__name__}", "ERROR")
            log(f"---------------------------------------------------------------\n")
            raise
        
        except Exception as e:
            log(f"❌ 未预期的错误 [{request_id}]: {e}", "ERROR")
            log(f"异常类型: {type(e).__name__}", "ERROR")
            import traceback
            log(f"堆栈跟踪:\n{traceback.format_exc()}", "ERROR")
            log(f"---------------------------------------------------------------\n")
            raise

    def call_rest_api(self, api: str, init_func) -> Generator:
        """Generic REST API call method with pagination support"""
        log(f"📄 调用分页API: {api}")
        url = self.BASE_URL + api
        page_count = 0
        
        while url:
            page_count += 1
            log(f"获取第 {page_count} 页数据: {url}")
            
            result = self._make_request('GET', url)
            data = result.json()
            
            items_count = len(data.get('value', []))
            log(f"第 {page_count} 页包含 {items_count} 个项目")
            
            for v in data.get('value', []):
                yield init_func(v, self)
            
            url = data.get('@odata.nextLink')
            if url:
                log(f"存在下一页: {url}")
            else:
                log(f"已到达最后一页")

    # ================= File operation methods =================

    def list_my_drive_items(self, path: str = '/', top: int = 100) -> Dict:
        """List files and folders in user's OneDrive"""
        log(f"📁 列出文件 - 路径: {path}, 数量: {top}")
        
        if path == '/' or path == '':
            url = f"{self.BASE_URL}/me/drive/root/children"
            log(f"使用根目录路径")
        else:
            clean_path = path.lstrip('/')
            url = f"{self.BASE_URL}/me/drive/root:/{clean_path}:/children"
            log(f"使用清理后的路径: {clean_path}")
        
        params = {'$top': min(top, 200)}
        log(f"查询参数: {params}")
        
        result = self._make_request('GET', url, params=params)
        response_data = result.json()
        
        items_count = len(response_data.get('value', []))
        log(f"✅ 成功获取 {items_count} 个项目")
        
        return response_data

    def get_my_drive_item(self, path: str = '/') -> Dict:
        """Get item information at specified path in user's OneDrive"""
        log(f"📄 获取项目信息 - 路径: {path}")
        
        if path == '/' or path == '':
            url = f"{self.BASE_URL}/me/drive/root"
            log(f"获取根目录信息")
        else:
            clean_path = path.lstrip('/')
            url = f"{self.BASE_URL}/me/drive/root:/{clean_path}"
            log(f"获取路径信息: {clean_path}")
        
        result = self._make_request('GET', url)
        item_data = result.json()
        
        log(f"✅ 成功获取项目: {item_data.get('name', 'Unknown')}")
        
        return item_data

    def url_to_base64(self, url: str) -> str:
        """Convert sharing URL to base64 encoding"""
        log(f"🔐 转换分享URL为base64: {url[:50]}...")
        
        encoded = base64.b64encode(url.encode())
        encoded = b'/shares/u!' + encoded.strip(b'=').replace(b'/',b'_').replace(b'+',b'-')
        result = encoded.decode()
        
        log(f"✅ Base64编码结果: {result[:50]}...")
        
        return result

    def get_driveitem(self, share_path: str):
        """Get drive item information"""
        log(f"🔗 获取共享驱动器项目: {share_path}")
        
        self.url_root = self.BASE_URL + self.url_to_base64(share_path) + '/driveItem'
        log(f"构建URL: {self.url_root}")
        
        result = self._make_request('GET', self.url_root)
        self.driveitem = result.json()
        self.root = self.BASE_URL + self.driveitem['parentReference']['path']
        
        log(f"✅ 成功获取驱动器项目，根路径: {self.root}")

    def listdir(self, path: str) -> Dir:
        """List directory contents"""
        log(f"📂 列出目录: {path}")
        
        result = self._make_request('GET', f"{self.root}{path}:/children")
        dir_data = Dir(result.json())
        
        log(f"✅ 目录内容: {dir_data}")
        
        return dir_data

    def downloadfile(self, file: str):
        """Download file"""
        log(f"⬇️ 下载文件: {file}")
        
        result = self._make_request('GET', f"{self.root}{file}:/content")
        filename = os.path.split(file)[1]
        
        log(f"文件名: {filename}, 大小: {len(result.content)} bytes")
        
        with open(filename, 'wb') as f:
            f.write(result.content)
        
        log(f"✅ 文件已保存: {filename}")

    def search_files(self, query: str) -> Dict:
        """Search files"""
        log(f"🔍 搜索文件: {query}")
        
        if not hasattr(self, 'root'):
            error_msg = "请先调用get_driveitem()设置根路径"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)
        
        drive_root = self.root.split('/root')[0]
        url = f"{drive_root}/root/search(q='{{{query}}}')"
        log(f"搜索URL: {url}")
        
        result = self._make_request('GET', url)
        search_results = result.json()
        
        results_count = len(search_results.get('value', []))
        log(f"✅ 找到 {results_count} 个结果")
        
        return search_results

    # ================= Email related methods =================
    
    def get_me_email(self) -> str:
        """Get current user's email address"""
        log(f"📧 获取用户邮箱")
        
        result = self._make_request('GET', f"{self.BASE_URL}/me/?$select=mail")
        email = result.json().get('mail')
        
        log(f"✅ 邮箱地址: {email}")
        
        return email

    def get_mail_with_filter(self, filter_func, folder: str = "inbox") -> Generator[Dict, None, None]:
        """Get emails with filter support"""
        log(f"📬 获取邮件 - 文件夹: {folder}")
        
        if folder:
            url = f"{self.BASE_URL}/me/mailFolders/{folder}/messages"
        else:
            url = f"{self.BASE_URL}/me/messages"
        
        params = filter_func() if callable(filter_func) else filter_func
        log(f"过滤参数: {params}")
        
        page_count = 0
        while url:
            page_count += 1
            log(f"获取第 {page_count} 页邮件")
            
            result = self._make_request('GET', url, params=params if page_count == 1 else None)
            data = result.json()
            
            mail_count = len(data.get('value', []))
            log(f"第 {page_count} 页包含 {mail_count} 封邮件")
            
            yield data
            
            url = data.get('@odata.nextLink')
            if url:
                log(f"存在下一页")
            else:
                log(f"已到达最后一页")

    def send_mail(self, to: List[str], cc: List[str], subject: str, body: str) -> Dict:
        """Send email"""
        log(f"📤 发送邮件")
        log(f"收件人: {to}")
        log(f"抄送: {cc}")
        log(f"主题: {subject}")
        log(f"内容长度: {len(body)} 字符")
        
        if not to:
            error_msg = "至少需要一个收件人"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)
        if not subject:
            error_msg = "主题不能为空"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)
        if not body:
            error_msg = "邮件内容不能为空"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)
        
        to_recipients = [{"emailAddress": {"address": addr}} for addr in to]
        cc_recipients = [{"emailAddress": {"address": addr}} for addr in cc] if cc else []
        
        content_type = "HTML" if any(tag in body.lower() for tag in ['<html>', '<div>', '<p>', '<br>', '<span>']) else "Text"
        log(f"检测到内容类型: {content_type}")
        
        message = {
            "message": {
                "subject": subject,
                "body": {
                    "contentType": content_type,
                    "content": body
                },
                "toRecipients": to_recipients,
                "ccRecipients": cc_recipients
            },
            "saveToSentItems": True
        }
        
        result = self._make_request('POST', f"{self.BASE_URL}/me/sendMail", json=message)
        
        log(f"✅ 邮件发送成功")
        
        return {"status": "sent", "status_code": result.status_code}

    def get_single_mail(self, message_id: str, select_fields: Optional[List[str]] = None) -> Dict:
        """Get single email"""
        log(f"📩 获取邮件详情 - ID: {message_id}")
        
        url = f"{self.BASE_URL}/me/messages/{message_id}"
        params = {}
        if select_fields:
            params['$select'] = ','.join(select_fields)
            log(f"选择字段: {select_fields}")
        
        result = self._make_request('GET', url, params=params)
        mail_data = result.json()
        
        log(f"✅ 邮件主题: {mail_data.get('subject', 'No subject')}")
        
        return mail_data

    def reply_to_mail(self, message_id: str, body: str, reply_all: bool = False) -> Dict:
        """Reply to email"""
        action = "全部回复" if reply_all else "回复"
        log(f"↩️ {action}邮件 - ID: {message_id}")
        log(f"回复内容长度: {len(body)} 字符")
        
        endpoint = 'replyAll' if reply_all else 'reply'
        url = f"{self.BASE_URL}/me/messages/{message_id}/{endpoint}"
        data = {"comment": body}
        
        result = self._make_request('POST', url, json=data)
        
        log(f"✅ {action}成功")
        
        return {"status": f"{'reply_all' if reply_all else 'reply'}_sent", "status_code": result.status_code}

    def forward_mail(self, message_id: str, to_recipients: List[str], cc_recipients: List[str] = None, body: Optional[str] = None) -> Dict:
        """Forward email"""
        log(f"➡️ 转发邮件 - ID: {message_id}")
        log(f"转发给: {to_recipients}")
        if cc_recipients:
            log(f"抄送: {cc_recipients}")
        if body:
            log(f"附加说明长度: {len(body)} 字符")
        
        url = f"{self.BASE_URL}/me/messages/{message_id}/forward"
        
        to_list = [{"emailAddress": {"address": addr}} for addr in to_recipients]
        data = {"toRecipients": to_list}
        
        if cc_recipients:
            cc_list = [{"emailAddress": {"address": addr}} for addr in cc_recipients]
            data["ccRecipients"] = cc_list

        if body:
            data["comment"] = body
        
        result = self._make_request('POST', url, json=data)
        
        log(f"✅ 转发成功")
        
        return {"status": "forwarded", "status_code": result.status_code}

    def get_mail_folders(self) -> Dict:
        """Get all email folders"""
        log(f"📁 获取邮件文件夹列表")
        
        result = self._make_request('GET', f"{self.BASE_URL}/me/mailFolders")
        folders_data = result.json()
        
        folder_count = len(folders_data.get('value', []))
        log(f"✅ 找到 {folder_count} 个文件夹")
        
        return folders_data

    def get_folder_messages(self, folder_id: str, filter_params=None) -> Generator[Dict, None, None]:
        """Get messages in folder"""
        log(f"📂 获取文件夹邮件 - 文件夹ID: {folder_id}")
        if filter_params:
            log(f"过滤参数: {filter_params}")
        
        url = f"{self.BASE_URL}/me/mailFolders/{folder_id}/messages"
        params = filter_params() if callable(filter_params) else (filter_params or {})
        
        page_count = 0
        while url:
            page_count += 1
            log(f"获取第 {page_count} 页")
            
            result = self._make_request('GET', url, params=params if page_count == 1 else None)
            data = result.json()
            
            mail_count = len(data.get('value', []))
            log(f"第 {page_count} 页包含 {mail_count} 封邮件")
            
            yield data
            
            url = data.get('@odata.nextLink')

    def get_mail_attachments(self, message_id: str) -> Dict:
        """Get email attachments"""
        log(f"📎 获取邮件附件列表 - 邮件ID: {message_id}")
        
        result = self._make_request('GET', f"{self.BASE_URL}/me/messages/{message_id}/attachments")
        attachments_data = result.json()
        
        attachment_count = len(attachments_data.get('value', []))
        log(f"✅ 找到 {attachment_count} 个附件")
        
        return attachments_data

    def download_attachment(self, message_id: str, attachment_id: str) -> Dict:
        """Download attachment"""
        log(f"⬇️ 下载附件 - 邮件ID: {message_id}, 附件ID: {attachment_id}")
        
        result = self._make_request('GET', f"{self.BASE_URL}/me/messages/{message_id}/attachments/{attachment_id}")
        attachment_data = result.json()
        
        log(f"✅ 附件名称: {attachment_data.get('name', 'Unknown')}")
        
        return attachment_data

    def search_mail(self, search_query: str, folder_id: Optional[str] = None) -> Dict:
        """Search emails"""
        log(f"🔍 搜索邮件 - 关键词: {search_query}")
        if folder_id:
            log(f"搜索文件夹: {folder_id}")
        
        if folder_id:
            url = f"{self.BASE_URL}/me/mailFolders/{folder_id}/messages"
        else:
            url = f"{self.BASE_URL}/me/messages"
        
        params = {"$search": f'"{search_query}"'}
        
        result = self._make_request('GET', url, params=params)
        total_result = result.json()
        
        # Handle pagination
        page_count = 1
        data = result.json()
        while '@odata.nextLink' in data:
            page_count += 1
            log(f"获取搜索结果第 {page_count} 页")
            
            url = data['@odata.nextLink']
            result = self._make_request('GET', url)
            data = result.json()
            total_result['value'].extend(data.get('value', []))
        
        total_result.pop('@odata.nextLink', None)
        
        total_count = len(total_result.get('value', []))
        log(f"✅ 搜索完成，共找到 {total_count} 封邮件")
        
        return total_result

    def get_single_mail_folder(self, folder_id: str) -> Dict:
        """Get single email folder information"""
        log(f"📁 获取文件夹信息 - ID: {folder_id}")
        
        result = self._make_request('GET', f"{self.BASE_URL}/me/mailFolders/{folder_id}")
        folder_data = result.json()
        
        log(f"✅ 文件夹名称: {folder_data.get('displayName', 'Unknown')}")
        
        return folder_data

    def save_each_mail_as_markdown(self, mail_data: Dict, saved_dir: str = 'mail'):
        """Save emails as Markdown files"""
        log(f"💾 保存邮件为Markdown - 目录: {saved_dir}")
        
        os.makedirs(saved_dir, exist_ok=True)
        
        saved_count = 0
        for mail in mail_data.get('value', []):
            fname = f'{time.time()} {mail["subject"]}'[:255]
            for schar in ("?", "/", ":", "*", "<", ">", "|", "\\", "\""): 
                fname = fname.replace(schar, "")
            
            markdown = HTML2Text().handle(mail["body"]["content"])
            filepath = f'{saved_dir}/{fname}.md'
            
            with open(filepath, 'w', encoding='utf8') as f:
                f.write(markdown)
            
            saved_count += 1
            log(f"已保存: {fname}.md")
        
        log(f"✅ 共保存 {saved_count} 封邮件")

    def save_attachments(self, message_id: str, save_dir: str = 'attachments') -> List[Dict]:
        """Save all attachments of an email"""
        log(f"💾 保存附件 - 邮件ID: {message_id}, 目录: {save_dir}")
        
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        
        attachments = self.get_mail_attachments(message_id)
        saved_files = []
        
        for attachment in attachments.get('value', []):
            if attachment.get('contentBytes'):
                filename = attachment.get('name', f"attachment_{attachment['id']}")
                filepath = os.path.join(save_dir, filename)
                
                content = base64.b64decode(attachment['contentBytes'])
                with open(filepath, 'wb') as f:
                    f.write(content)
                
                file_info = {
                    "filename": filename,
                    "filepath": filepath,
                    "size": len(content),
                    "content_type": attachment.get('contentType', 'unknown')
                }
                saved_files.append(file_info)
                
                log(f"已保存附件: {filename} ({len(content)} bytes)")
        
        log(f"✅ 共保存 {len(saved_files)} 个附件")
        
        return saved_files

    def get_unread_count(self, folder_id: Optional[str] = None) -> int:
        """Get unread email count"""
        log(f"📊 获取未读邮件数量")
        
        if folder_id:
            log(f"指定文件夹: {folder_id}")
            folder_info = self.get_single_mail_folder(folder_id)
        else:
            log(f"获取收件箱")
            folders = self.get_mail_folders()
            inbox = next((f for f in folders.get('value', []) if f['displayName'] == 'Inbox'), None)
            folder_info = inbox
        
        unread_count = folder_info.get('unreadItemCount', 0) if folder_info else 0
        log(f"✅ 未读邮件数: {unread_count}")
        
        return unread_count

    # ================= OneNote related methods =================
    
    def get_notebooks(self, user_email: Optional[str] = None) -> Generator:
        """Get notebooks"""
        log(f"📓 获取笔记本列表")
        if user_email:
            log(f"指定用户: {user_email}")
        
        if user_email: 
            url = f'/users/{user_email}/onenote/notebooks'
        else: 
            url = '/me/onenote/notebooks'
        
        return self.call_rest_api(url, lambda v, o: v)

    def get_sections(self, notebook_id: str, user_email: Optional[str] = None, top: int = 100) -> Generator:
        """
        Get notebook sections (smart method selection, automatically handles large datasets)
        
        Prioritizes direct retrieval method, if encountering 403 error (dataset too large),
        automatically switches to filtered query method (returns recently modified sections)
        
        Args:
            notebook_id: Notebook ID
            user_email: Optional user email
            top: Number of items to return per request (default 100)
        
        Yields:
            Section data dictionaries
        """
        log(f"📑 获取章节 - 笔记本ID: {notebook_id}")
        
        if user_email:
            base_url = f'/users/{user_email}/onenote/notebooks'
            sections_base = f'/users/{user_email}/onenote/sections'
        else:
            base_url = '/me/onenote/notebooks'
            sections_base = '/me/onenote/sections'
        
        # Method 1: Direct retrieval (suitable for small datasets)
        direct_url = f'{base_url}/{notebook_id}/sections?$top={top}'
        log(f"🔄 尝试方法1: 直接获取章节")
        log(f"请求URL: {direct_url}")
        
        try:
            # Try direct retrieval
            full_url = f"{self.BASE_URL}{direct_url}"
            result = self._make_request('GET', full_url)
            data = result.json()
            
            items_count = len(data.get('value', []))
            log(f"✅ 方法1成功！获取到 {items_count} 个章节")
            
            # Use call_rest_api to handle pagination
            for v in data.get('value', []):
                yield v
            
            # Handle subsequent pagination
            next_url = data.get('@odata.nextLink')
            while next_url:
                log(f"📄 获取下一页: {next_url}")
                result = self._make_request('GET', next_url)
                data = result.json()
                
                for v in data.get('value', []):
                    yield v
                
                next_url = data.get('@odata.nextLink')
            
            return
            
        except requests.HTTPError as e:
            if e.response.status_code == 403:
                # 403 error, try alternative method
                log(f"⚠️ 方法1失败 (403 Forbidden) - 可能数据量超过5000项", "WARN")
                log(f"错误详情: {e.response.text[:200]}", "WARN")
                log(f"🔄 自动切换到方法2: 过滤查询（最近修改）", "WARN")
                
                try:
                    # Method 2: Filtered query (bypasses 5000 item limit)
                    filter_url = f"{sections_base}?$filter=parentNotebook/id eq '{notebook_id}'&$top={min(top, 100)}&$orderby=lastModifiedDateTime desc"
                    log(f"请求URL: {filter_url}")
                    
                    full_url = f"{self.BASE_URL}{filter_url}"
                    result = self._make_request('GET', full_url)
                    data = result.json()
                    
                    items_count = len(data.get('value', []))
                    log(f"✅ 方法2成功！获取到 {items_count} 个章节（按最近修改排序）", "SUCCESS")
                    log(f"📌 注意: 由于数据量限制，仅返回最近修改的章节", "INFO")
                    
                    # Handle data and pagination
                    for v in data.get('value', []):
                        yield v
                    
                    next_url = data.get('@odata.nextLink')
                    page_count = 1
                    while next_url:
                        page_count += 1
                        log(f"📄 获取第 {page_count} 页")
                        result = self._make_request('GET', next_url)
                        data = result.json()
                        
                        for v in data.get('value', []):
                            yield v
                        
                        next_url = data.get('@odata.nextLink')
                    
                    return
                    
                except Exception as filter_error:
                    log(f"❌ 方法2也失败: {filter_error}", "ERROR")
                    raise Exception(f"无法获取章节 - 所有方法都失败。原始错误: {str(e)}, 备用方法错误: {str(filter_error)}")
            else:
                # Other HTTP errors
                log(f"❌ 方法1失败 (HTTP {e.response.status_code}): {e}", "ERROR")
                raise
                
        except Exception as e:
            log(f"❌ 获取章节时发生未预期错误: {e}", "ERROR")
            raise

    def get_pages(self, section_id: str, user_email: Optional[str] = None) -> Generator:
        """Get section pages"""
        log(f"📄 获取页面 - 章节ID: {section_id}")
        
        if user_email: 
            url = f'/users/{user_email}/onenote/sections'
        else: 
            url = f'/me/onenote/sections'
        
        return self.call_rest_api(f'{url}/{section_id}/pages', lambda v, o: v)

    def get_page_content(self, page_id: str, user_email: Optional[str] = None) -> bytes:
        """Get page content"""
        log(f"📝 获取页面内容 - 页面ID: {page_id}")
        
        if user_email: 
            url = f'/users/{user_email}/onenote/pages'
        else: 
            url = '/me/onenote/pages'
        
        full_url = f"{self.BASE_URL}{url}/{page_id}/content"
        result = self._make_request('GET', full_url)
        
        content_size = len(result.content)
        log(f"✅ 页面内容大小: {content_size} bytes")
        
        return result.content

    # ================= Teams related methods =================
    
    def get_chats(self) -> Generator[Chat, None, None]:
        """Get Teams chats"""
        log(f"💬 获取Teams聊天列表")
        return self.call_rest_api('/me/chats', lambda v, o: Chat(v, o))

    def get_chat_messages(self, chat_id: str) -> Generator[Message, None, None]:
        """Get chat messages"""
        log(f"💬 获取聊天消息 - 聊天ID: {chat_id}")
        return self.call_rest_api(f'/chats/{chat_id}/messages', lambda v, o: Message(v, o))

    def send_chat_message(self, chat_id: str, content: str, content_type: str = "text") -> Dict[str, Any]:
        """
        Send a message to a Teams chat
        
        Args:
            chat_id: Chat ID to send message to
            content: Message content
            content_type: Content type - "text" or "html" (default: "text")
            
        Returns:
            Dictionary containing sent message data
        """
        log(f"📤 发送聊天消息 - 聊天ID: {chat_id}")
        log(f"内容类型: {content_type}, 内容长度: {len(content)} 字符")
        
        if not content or not content.strip():
            error_msg = "消息内容不能为空"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)
        
        url = f"{self.BASE_URL}/chats/{chat_id}/messages"
        
        message_data = {
            "body": {
                "contentType": content_type,
                "content": content
            }
        }
        
        result = self._make_request('POST', url, json=message_data)
        response_data = result.json()
        
        log(f"✅ 消息发送成功 - 消息ID: {response_data.get('id', 'Unknown')}")
        
        return response_data

    def create_chat(self, chat_type: str, members: List[Dict[str, str]], topic: Optional[str] = None) -> Dict[str, Any]:
        """
        Create a new Teams chat
        
        Args:
            chat_type: Chat type - "oneOnOne" or "group"
            members: List of member dictionaries with 'user_id' and optional 'roles'
                    Example: [{"user_id": "user@example.com", "roles": ["owner"]}]
            topic: Chat topic (required for group chats, optional for oneOnOne)
            
        Returns:
            Dictionary containing created chat data
        """
        log(f"➕ 创建聊天 - 类型: {chat_type}")
        if topic:
            log(f"主题: {topic}")
        log(f"成员数量: {len(members)}")
        
        if chat_type not in ["oneOnOne", "group"]:
            error_msg = f"不支持的聊天类型: {chat_type}。仅支持 'oneOnOne' 或 'group'"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)
        
        if chat_type == "group" and not topic:
            error_msg = "群聊必须提供主题"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)
        
        if not members or len(members) < 1:
            error_msg = "至少需要一个成员"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)

        # ── 兜底：获取当前用户，确保 caller 在 members 中 ──────────────────
        try:
            me_response = self._make_request('GET', f"{self.BASE_URL}/me")
            me_data = me_response.json()
            current_user_id = me_data.get('id', '')
            current_user_upn = me_data.get('userPrincipalName', '')
            log(f"🔍 当前用户: {current_user_upn} ({current_user_id})")
        except Exception as e:
            log(f"⚠️ 无法获取当前用户信息，跳过兜底检查: {e}", "WARN")
            current_user_id = ''
            current_user_upn = ''

        if current_user_id or current_user_upn:
            # 检查 caller 是否已在 members 列表中（匹配 id 或 email/upn）
            member_ids = [m.get('user_id', '').lower() for m in members]
            caller_in_members = (
                current_user_id.lower() in member_ids or
                current_user_upn.lower() in member_ids
            )

            if not caller_in_members:
                log(f"⚠️ 当前用户不在 members 列表中，自动补充: {current_user_upn}", "WARN")
                members = [{"user_id": current_user_id, "roles": ["owner"]}] + list(members)
                log(f"✅ 已将当前用户插入 members 首位，新成员数量: {len(members)}")
        # ────────────────────────────────────────────────────────────────────

        url = f"{self.BASE_URL}/chats"
        
        # Build member list
        chat_members = []
        for member in members:
            user_id = member.get('user_id')
            if not user_id:
                error_msg = "每个成员必须包含 'user_id'"
                log(f"❌ {error_msg}", "ERROR")
                raise ValueError(error_msg)
            
            member_data = {
                "@odata.type": "#microsoft.graph.aadUserConversationMember",
                "roles": member.get('roles', ["owner"]),
                "user@odata.bind": f"https://graph.microsoft.com/v1.0/users('{user_id}')"
            }
            chat_members.append(member_data)
        
        chat_data = {
            "chatType": chat_type,
            "members": chat_members
        }
        
        if topic:
            chat_data["topic"] = topic
        
        result = self._make_request('POST', url, json=chat_data)
        response_data = result.json()
        
        log(f"✅ 聊天创建成功 - 聊天ID: {response_data.get('id', 'Unknown')}")
        
        return response_data

    def update_chat_topic(self, chat_id: str, topic: str) -> Dict[str, Any]:
        """
        Update chat topic (group chats only)
        
        Args:
            chat_id: Chat ID to update
            topic: New topic for the chat
            
        Returns:
            Dictionary containing update status
        """
        log(f"✏️ 更新聊天主题 - 聊天ID: {chat_id}")
        log(f"新主题: {topic}")
        
        if not topic or not topic.strip():
            error_msg = "主题不能为空"
            log(f"❌ {error_msg}", "ERROR")
            raise ValueError(error_msg)
        
        url = f"{self.BASE_URL}/chats/{chat_id}"
        
        update_data = {
            "topic": topic
        }
        
        result = self._make_request('PATCH', url, json=update_data)
        
        log(f"✅ 聊天主题更新成功")
        
        return {"status": "updated", "status_code": result.status_code, "chat_id": chat_id, "new_topic": topic}


    def get_file_content(self, path: str, max_size_mb: float = 1.0) -> Dict[str, Any]:
        """
        Get file content
        
        Supported file types:
        - Text files: Return content directly
        - xlsx files: Convert to CSV format and return all worksheets
        - Other files: Return download link
        
        Args:
            path: File path
            max_size_mb: Maximum size limit for text files (MB), xlsx files fixed at 5MB
            
        Returns:
            Dictionary containing file content or download link
        """
        try:
            # Get file information
            item_info = self.get_my_drive_item(path)
            
            # Check if it's a file
            if 'file' not in item_info:
                return {
                    "success": False,
                    "error": f"Path '{path}' is not a file"
                }
            
            file_name = item_info.get('name', 'unknown')
            file_size = item_info.get('size', 0)
            size_mb = file_size / (1024 * 1024)
            mime_type = item_info.get('file', {}).get('mimeType', '')
            
            # Get download URL
            download_url = item_info.get('@microsoft.graph.downloadUrl')
            if not download_url:
                return {
                    "success": False,
                    "error": "No download URL available for this file"
                }
            
            # Check if it's an Excel file
            is_xlsx = (
                file_name.lower().endswith('.xlsx') or 
                'spreadsheet' in mime_type.lower() or
                'excel' in mime_type.lower()
            )
            
            if is_xlsx:
                # Excel file: 5MB limit, convert to CSV
                if size_mb > 5.0:
                    return {
                        "type": "xlsx_too_large",
                        "name": file_name,
                        "size": file_size,
                        "size_mb": round(size_mb, 2),
                        "download_link": download_url,
                        "message": f"Excel file is too large ({size_mb:.2f}MB). Maximum allowed: 5MB. Please use the download link."
                    }
                
                # Download file content
                content_response = self._make_request('GET', download_url)
                content = content_response.content
                
                # Process xlsx file
                return self._process_xlsx_to_csv(content, file_name, max_size_mb=5.0)
            
            # Determine if it's a text file
            text_extensions = {
                '.txt', '.md', '.csv', '.json', '.xml', '.yaml', '.yml',
                '.py', '.js', '.ts', '.java', '.cpp', '.c', '.h', '.cs', '.go',
                '.html', '.css', '.scss', '.less', '.sh', '.bash', '.bat',
                '.log', '.conf', '.config', '.ini', '.env', '.gitignore',
                '.sql', '.r', '.rb', '.php', '.swift', '.kt', '.rs', '.dart',
                '.vue', '.jsx', '.tsx', '.svelte', '.astro'
            }
            
            file_ext = '.' + file_name.split('.')[-1].lower() if '.' in file_name else ''
            is_text = file_ext in text_extensions or 'text' in mime_type.lower()
            
            if not is_text:
                # Binary file: return download link
                return {
                    "type": "binary",
                    "name": file_name,
                    "size": file_size,
                    "size_mb": round(size_mb, 2),
                    "mime_type": mime_type,
                    "download_link": download_url,
                    "message": "This is a binary file. Please use the download link to access it."
                }
            
            # Text file processing
            if size_mb > max_size_mb:
                return {
                    "type": "text_too_large",
                    "name": file_name,
                    "size": file_size,
                    "size_mb": round(size_mb, 2),
                    "download_link": download_url,
                    "message": f"Text file is too large ({size_mb:.2f}MB). Maximum allowed: {max_size_mb}MB. Please use the download link."
                }
            
            # Download text file content
            content_response = self._make_request('GET', download_url)
            content = content_response.content
            
            # Try to decode text
            encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'big5', 'latin1']
            text_content = None
            used_encoding = None
            
            for encoding in encodings:
                try:
                    text_content = content.decode(encoding)
                    used_encoding = encoding
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            
            if text_content is None:
                return {
                    "type": "decode_error",
                    "name": file_name,
                    "size": file_size,
                    "size_mb": round(size_mb, 2),
                    "download_link": download_url,
                    "message": "Unable to decode file as text. Please use the download link."
                }
            
            return {
                "type": "text",
                "name": file_name,
                "size": file_size,
                "size_mb": round(size_mb, 2),
                "content": text_content,
                "encoding": used_encoding,
                "char_count": len(text_content),
                "line_count": text_content.count('\n') + 1
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

    def get_download_link(self, path: str) -> Dict[str, Any]:
        """
        Get file download link
        
        Args:
            path: File path
            
        Returns:
            Dictionary containing download link
        """
        log(f"🔗 获取下载链接 - 路径: {path}")
        
        try:
            item_info = self.get_my_drive_item(path)
            
            # Check if it's a file
            if 'folder' in item_info:
                error_msg = f"路径 '{path}' 是文件夹，不是文件"
                log(f"❌ {error_msg}", "ERROR")
                return {
                    "success": False,
                    "error": error_msg
                }
            
            file_name = item_info.get('name', 'unknown')
            file_size = item_info.get('size', 0)
            download_url = item_info.get('@microsoft.graph.downloadUrl')
            
            if not download_url:
                error_msg = "无法获取下载链接"
                log(f"❌ {error_msg}", "ERROR")
                return {
                    "success": False,
                    "error": error_msg
                }
            
            log(f"✅ 下载链接获取成功: {file_name}")
            
            return {
                "success": True,
                "name": file_name,
                "size": file_size,
                "size_mb": round(file_size / (1024 * 1024), 2),
                "download_link": download_url
            }
            
        except Exception as e:
            log(f"❌ 获取下载链接失败: {e}", "ERROR")
            return {
                "success": False,
                "error": str(e)
            }
    def _process_xlsx_to_csv(self, content: bytes, filename: str, max_size_mb: float = 5.0) -> Dict[str, Any]:
        """
        Convert xlsx file content to CSV format
        
        Args:
            content: Binary content of xlsx file
            filename: File name
            max_size_mb: Maximum processing size (MB)
            
        Returns:
            Dictionary containing CSV data for all worksheets
        """
        import csv
        
        try:
            # Check file size
            size_mb = len(content) / (1024 * 1024)
            if size_mb > max_size_mb:
                return {
                    "success": False,
                    "type": "xlsx_too_large",
                    "name": filename,
                    "size": len(content),
                    "size_mb": round(size_mb, 2),
                    "error": f"Excel file too large ({size_mb:.2f}MB). Maximum allowed: {max_size_mb}MB"
                }
            
            # Load workbook
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            
            sheets_data = {}
            total_rows = 0
            total_cells = 0
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Convert to CSV format
                csv_output = io.StringIO()
                csv_writer = csv.writer(csv_output)
                
                row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # Convert None to empty string, other values to string
                    cleaned_row = [str(cell) if cell is not None else '' for cell in row]
                    csv_writer.writerow(cleaned_row)
                    row_count += 1
                    total_cells += len(cleaned_row)
                
                csv_content = csv_output.getvalue()
                csv_output.close()
                
                # Only save non-empty worksheets
                if row_count > 0:
                    sheets_data[sheet_name] = csv_content
                    total_rows += row_count
            
            workbook.close()
            
            if not sheets_data:
                return {
                    "success": False,
                    "type": "xlsx_empty",
                    "name": filename,
                    "error": "Excel file contains no data"
                }
            
            return {
                "success": True,
                "type": "xlsx",
                "name": filename,
                "size": len(content),
                "size_mb": round(size_mb, 2),
                "sheets": sheets_data,
                "sheet_count": len(sheets_data),
                "total_rows": total_rows,
                "total_cells": total_cells,
                "sheet_names": list(sheets_data.keys())
            }
            
        except Exception as e:
            return {
                "success": False,
                "type": "xlsx_parse_error",
                "name": filename,
                "error": f"Failed to parse Excel file: {str(e)}"
            }

# Convenient factory function
async def create_onedrive_service(token: str) -> OneDriverService:
    """
    Create and authenticate OneDrive service instance
    
    Args:
        token: unique_token
        
    Returns:
        Authenticated OneDriverService instance
    """
    log(f"🏭 工厂函数: 创建 OneDrive 服务")
    log(f"Token: {token[:20]}...{token[-10:]}")
    
    service = OneDriverService(token)
    
    log(f"✅ 服务创建完成")
    
    return service